from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib.auth.models import User
from django.utils import timezone
from django.core.exceptions import FieldError

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from openpyxl import load_workbook

from main.models import EnterProduct, Product
from main import models
from main import models
from .funcs import search_with_fields, pagenator_page, search_with_fields_product



def dashboard(request):
    categorys = models.Category.objects.all()
    # products = models.Product.objects.filter(is_active=True)
    # users = User.objects.filter(is_satff=False)
    context = {
        'categorys':categorys,
        # 'products':products,
        # 'users':users,
    }
    return render(request, 'dashboard/index.html', context)

def create_category(request):
    if request.method == 'POST':
        models.Category.objects.create(
            name=request.POST['name']
        )
        return redirect('dashboard:category')
    return render(request, 'dashboard/category/create.html')


def category_list(request):
    categorys = models.Category.objects.all()
    return render(request, 'dashboard/category/list.html', {'categorys':categorys})


def category_detail(request, id):
    category = models.Category.objects.get(id=id)
    products = models.Product.objects.filter(category=category, is_active=True)
    context = {
        'category':category,
        'products':products
    }
    return render(request, 'dashboard/category/list.html', context)


def category_update(request, id):
    category = models.Category.objects.get(id=id)
    if request.method == 'POST':
        category.name = request.POST['name']
        category.save()
        return redirect('dashboard:category')
    return render(request, 'dashboard/category/update.html', {'category':category})

def category_delete(request, id):
    models.Category.objects.get(id=id).delete()
    return redirect('dashboard:category')


# Product
def product_create(request):

    if request.method == 'POST':
        name = request.POST['name']
        description = request.POST['description']
        currency = request.POST['currency']
        quantity = request.POST['quantity']
        price = request.POST['price']
        # discount_price = request.POST['discount_price']
        category = models.Category.objects.get(id=request.POST['category_id'])
        baner_image = request.FILES['baner_image']
        models.Product.objects.create(
            name=name,
            description = description,
            category = category,
            currency = currency,
            quantity = quantity,
            price = price,
            # discount_price = discount_price,
            baner_image = baner_image
        )
        return redirect('dashboard:product')
    context = {
        'categorys':models.Category.objects.all(),
    }
    return render(request, 'dashboard/product/create.html', context)


def product(request):
    result = search_with_fields_product(request)
    try: 
        product = models.Product.objects.filter(**result)
    except FieldError as err:
        del result[err.__doc__.split()[3][1:-1]]
        product = models.Product.objects.filter(**result)
        
    context = {'product': pagenator_page(product, 10, request)}
    return render(request, 'dashboard/product/list.html', context)


def product_update(request, id):
    product = models.Product.objects.get(id=id)
    if request.method == 'POST':
        category = models.Category.objects.get(id=request.POST['category_id'])
        product.name = request.POST['name']
        product.description = request.POST['description']
        product.quantity = request.POST['quantity']
        product.price = request.POST['price']
        product.currency = request.POST['currency']
        # product.discount_price = request.POST['discount_price']
        product.category=category
        baner_image = request.FILES.get('product_image')
        if baner_image:
            product.baner_image = baner_image
        product.save()

    context = {
        'product':product,
        'categorys':models.Category.objects.all(),
    }
    return render(request, 'dashboard/product/update.html', context)


def product_delete(request, id):
    models.Product.objects.get(id=id).delete()
    return redirect('dashboard:product')


def create_enter(request):
    if request.method == 'POST':
        product_id = request.POST['product_id']
        quantity = int(request.POST['quantity'])
        models.EnterProduct.objects.create(
            product_id=product_id,
            quantity=quantity
        )
        return redirect('dashboard:list_enter')
    return render(request, 'dashboard/enter/create.html', {'products':models.Product.objects.all()})


def update_enter(request, id):
    if request.method == 'POST':
        quantity = int(request.POST['quantity'])
        enter = models.EnterProduct.objects.get(id=id)
        enter.quantity = quantity
        enter.save()
    return redirect('dashboard:list_enter')


def delete_enter(request, id):
    models.EnterProduct.objects.get(id=id).delete()
    return redirect('dashboard:list_enter') 


def list_enter(request):
    result = search_with_fields(request)
    print(result)
    try: 
        enters = models.EnterProduct.objects.filter(**result)
    except FieldError as err:
        del result['page']

        enters = models.EnterProduct.objects.filter(**result)
        
    context = {'enters': pagenator_page(enters, 10, request)}
    return render(request, 'dashboard/enter/list.html', context)


def print_to_excel(request):
    enters = models.EnterProduct.objects.all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Enter Products"
    headers = ['Product', 'Date', 'Quantity']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
    date_style = NamedStyle(name="datetime", number_format="YYYY-MM-DD HH:mm:ss")
    ws.cell(row=1, column=2).style = date_style  
    
    for row_num, enter in enumerate(enters, 2):
        ws.cell(row=row_num, column=1).value = enter.product.name 
        local_datetime = timezone.localtime(enter.created_at)
        ws.cell(row=row_num, column=2).value = local_datetime.replace(tzinfo=None)
        ws.cell(row=row_num, column=3).value = enter.quantity
        
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=enter_products.xlsx'
    wb.save(response)
    return response

def print_to_excel_project(request):
    results = models.Product.objects.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Enter Products"
    headers = ['Name', 'Description', 'Quantity', 'Price', 'Currency']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
    for row_num, result in enumerate(results, 2):
        ws.cell(row=row_num, column=1).value = result.name
        ws.cell(row=row_num, column=2).value = result.description
        ws.cell(row=row_num, column=3).value = result.quantity
        ws.cell(row=row_num, column=4).value = result.price
        ws.cell(row=row_num, column=5).value = 'So`m'
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=enter_products.xlsx'
    wb.save(response)
    print(response)
    return response


def product_detail(request, id):
    product = models.Product.objects.get(id=id)
    enters = models.EnterProduct.objects.filter(product=product)
    products = models.Product.objects.get(id=id)
    recomendation = models.Product.objects.filter(
        category_id=product.category.id).exclude(id=product.id)[:3]
    categorys = models.Category.objects.all()
    images = models.Product.objects.filter(baner_image=product)

    data = {
        'product': product,
        'enters': enters,
        'images':images,
        'categorys':categorys,
        'products':products,
        'recomendation':recomendation
    }
    return render(request, 'dashboard/product/product_detail.html', data)

def upload_and_display(request):
    if request.method == 'POST':
        fayl = request.FILES['fayl']
        workbook = load_workbook(fayl)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            product_name, date, quantity = row

            if quantity is None:
                continue
            product, created = Product.objects.get_or_create(name=product_name)

            EnterProduct.objects.create(
                product=product,
                quantity=quantity,
                created_at=date
            )
            return redirect('dashboard:list_enter') 
    return render(request, 'dashboard/enter/excel_enter.html')